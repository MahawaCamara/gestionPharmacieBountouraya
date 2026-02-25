#### Mes importations ####
from decimal import Decimal
from django.conf import settings
from django.urls import reverse
import openpyxl
from django.db.models.functions import TruncMonth
import json
from django.core.mail import send_mail
from django.utils.timezone import now
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from administrateur.forms import FormForm
from pharmacien.models import Pharmacy
from Produit.models import ProductAvailability, Form, Product
from django.contrib.auth import get_user_model
from abonnement.models import Abonnement
from django.utils import timezone
from django.contrib import messages
from django.db.models import Count, F, ExpressionWrapper, DateField, Sum
from django.db.models.functions import TruncWeek
from datetime import timedelta
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.html import strip_tags

User = get_user_model()

def admin_required(view_func):
    return user_passes_test(lambda u: u.is_superuser)(view_func)

#### Vue pour gerer la pagination ####
def paginate(request, queryset, per_page=5):
    """Gestion de la pagination pour les objets"""
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return page_obj

####### Vue pour afficher le dashboard admin ###########
@login_required
@admin_required
def dashboard(request):
    # Comptage des éléments
    pharmacies = Pharmacy.objects.count()
    produits = ProductAvailability.objects.count()  # ✅ correction ici
    utilisateurs = User.objects.count()
    abonnements = Abonnement.objects.count()

    # Récupérer la date actuelle
    now = timezone.now()
    week_start = now - timedelta(weeks=1)

    # Données pharmacies par semaine
    pharmacy_data = (
        Pharmacy.objects.filter(created_at__gte=week_start)
        .annotate(week=TruncWeek('created_at'))
        .values('week')
        .annotate(count=Count('id'))
        .order_by('week')
    )

    # ✅ Données produits à partir de ProductAvailability, pas Product
    product_data = (
        ProductAvailability.objects.filter(created_at__gte=week_start)
        .annotate(week=TruncWeek('created_at'))
        .values('week')
        .annotate(count=Count('id'))
        .order_by('week')
    )

    # Données abonnements par semaine
    subscription_data = (
        Abonnement.objects.filter(date_debut__gte=week_start)
        .annotate(week=TruncWeek('date_debut'))
        .values('week')
        .annotate(count=Count('id'))
        .order_by('week')
    )

    def prepare_chart_data(data):
        labels = [entry['week'].strftime('%Y-%m-%d') for entry in data]
        counts = [entry['count'] for entry in data]
        return {'labels': labels, 'data': counts}

    pharmacy_data = prepare_chart_data(pharmacy_data)
    product_data = prepare_chart_data(product_data)
    subscription_data = prepare_chart_data(subscription_data)

    return render(request, "admin/dashboard.html", {
        "pharmacies": pharmacies,
        "produits": produits,  # ✅ déjà modifié
        "utilisateurs": utilisateurs,
        "abonnements": abonnements,
        "pharmacy_data": pharmacy_data,
        "product_data": product_data,
        "subscription_data": subscription_data,
    })

# Vue d'envoie d'emails aux users bloqués
def send_pharmacy_notification(pharmacy, subject, html_message):
    recipient = pharmacy.email
    plain_message = strip_tags(html_message)
    print(f"[ENVOI EMAIL] Destinataire : {recipient}")  # Debug

    send_mail(
        subject=subject,
        message=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[recipient],
        html_message=html_message,
        fail_silently=False
    )
    
###### Vue pour lister toutes les pharmacies ######    
@login_required
@admin_required
def manage_pharmacies(request):
    search_term = request.GET.get('search', '')
    filter_by = request.GET.get('filter_by', 'pharmacy_name')
    status_term = request.GET.get('status', '')
    
    # Appliquer le filtre selon le champ sélectionné
    if filter_by == "pharmacy_name":
        pharmacies = Pharmacy.objects.filter(pharmacy_name__icontains=search_term)
    elif filter_by == "address":
        pharmacies = Pharmacy.objects.filter(address__icontains=search_term)
    elif filter_by == "status":
        if status_term == "approved":
            pharmacies = Pharmacy.objects.filter(is_approved=True)
        elif status_term == "not_approved":
            pharmacies = Pharmacy.objects.filter(is_approved=False)
        else:
            pharmacies = Pharmacy.objects.all()

    pharmacies = pharmacies.select_related("user")
    
    # Pagination
    page_obj = paginate(request, pharmacies, 5)
    return render(request, "admin/pharmaciens.html", {"page_obj": page_obj, "search_term": search_term, "filter_by": filter_by, "status_term": status_term})

############ Vue pour approuver une pharmacie ############
@login_required
@admin_required
def approve_pharmacy(request, id):
    pharmacy = get_object_or_404(Pharmacy, id=id)
    pharmacy.is_approved = not pharmacy.is_approved
    pharmacy.save()

    if pharmacy.email:
        if pharmacy.is_approved:
            subject = "Pharmacie approuvée - PharmaConnect"
            html_message = f"""
            <div style="font-family: Arial; padding: 20px;">
                <h2 style="color: #28a745;">Pharmacie approuvée</h2>
                <p>Bonjour <strong>{pharmacy.pharmacy_name}</strong>,</p>
                <p>Nous sommes heureux de vous informer que votre pharmacie a été <strong>approuvée</strong> sur la plateforme <strong>PharmaConnect</strong>.</p>
                <p>Vous pouvez désormais publier vos produits et bénéficier de tous les services offerts.</p>
                <p>Merci de votre confiance.</p>
                <p>Cordialement,<br><strong>L'équipe PharmaConnect</strong></p>
            </div>
            """
            messages.success(request, "La pharmacie a été approuvée avec succès.")
        else:
            subject = "Pharmacie désapprouvée - PharmaConnect"
            html_message = f"""
            <div style="font-family: Arial; padding: 20px;">
                <h2 style="color: #dc3545;">Pharmacie désapprouvée</h2>
                <p>Bonjour <strong>{pharmacy.pharmacy_name}</strong>,</p>
                <p>Votre pharmacie a été temporairement <strong>désapprouvée</strong> par notre équipe.</p>
                <p>Les motifs possibles sont :</p>
                <ul>
                    <li>Informations manquantes ou incorrectes</li>
                    <li>Problèmes de conformité</li>
                </ul>
                <p>Merci de vérifier et compléter votre profil si nécessaire.</p>
                <p>Pour toute question, contactez-nous à <a href="mailto:pharmaconnect200@gmail.com">pharmaconnect200@gmail.com</a>.</p>
                <p>Cordialement,<br><strong>L'équipe PharmaConnect</strong></p>
            </div>
            """
            messages.warning(request, "La pharmacie a été désapprouvée.")

        send_pharmacy_notification(pharmacy, subject, html_message)

    return redirect('administration:manage_pharmacies')

############ Vue pour supprimer une pharmacie ############
@login_required
@admin_required
def delete_pharmacy(request, id):
    pharmacy = get_object_or_404(Pharmacy, id=id)
    user = pharmacy.user
    email = pharmacy.email
    name = pharmacy.pharmacy_name

    # Email avant suppression
    subject = "Suppression de votre pharmacie - PharmaConnect"
    html_message = f"""
    <div style="font-family: Arial; padding: 20px;">
        <h2 style="color: #dc3545;">Pharmacie supprimée</h2>
        <p>Bonjour <strong>{name}</strong>,</p>
        <p>Votre pharmacie a été <strong style="color:red;">supprimée</strong> de notre plateforme <strong>PharmaConnect</strong>.</p>
        <p>Raisons possibles :</p>
        <ul>
            <li>Données inexactes ou non vérifiées</li>
            <li>Produits non conformes</li>
            <li>Violation des conditions d'utilisation</li>
        </ul>
        <p>Pour toute question, contactez : <a href="mailto:pharmaconnect200@gmail.com">pharmaconnect200@gmail.com</a>.</p>
        <p><strong>L'équipe PharmaConnect</strong></p>
    </div>
    """
    send_pharmacy_notification(pharmacy, subject, html_message)

    # Supprimer produits puis compte
    Product.objects.filter(created_by=pharmacy).delete()
    pharmacy.delete()
    user.delete()

    messages.success(request, "Pharmacie et ses produits supprimés avec succès.")
    return redirect("administration:manage_pharmacies")

############ Vue pour export les données de la pharmacie en excel ############
@login_required
@admin_required
def export_pharmacies_excel(request):
    # Récupérer les filtres comme dans manage_pharmacies pour exporter la même liste
    search_term = request.GET.get('search', '')
    filter_by = request.GET.get('filter_by', 'pharmacy_name')
    status_term = request.GET.get('status', '')

    if filter_by == "pharmacy_name":
        pharmacies = Pharmacy.objects.filter(pharmacy_name__icontains=search_term)
    elif filter_by == "address":
        pharmacies = Pharmacy.objects.filter(address__icontains=search_term)
    elif filter_by == "status":
        if status_term == "approved":
            pharmacies = Pharmacy.objects.filter(is_approved=True)
        elif status_term == "not_approved":
            pharmacies = Pharmacy.objects.filter(is_approved=False)
        else:
            pharmacies = Pharmacy.objects.all()
    else:
        pharmacies = Pharmacy.objects.all()

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pharmacies"

    # Écrire l'en-tête
    headers = ["Nom", "Adresse", "Téléphone", "Email", "Statut"]
    ws.append(headers)

    # Écrire les données
    for p in pharmacies:
        statut = "Approuvée" if p.is_approved else "Non approuvée"
        ws.append([p.pharmacy_name, p.address, p.phone_number, p.email, statut])

    # Ajuster la largeur des colonnes (optionnel)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="pharmacies.xlsx"'

    wb.save(response)
    return response

# Gestions des produits
############ Vue pour lister tous les produits ############
@login_required
@admin_required
def manage_products(request):
    search_term = request.GET.get('search', '')
    filter_by = request.GET.get('filter_by', '')

    # Recherche sur ProductAvailability, joint avec Product
    query = Q()
    if search_term:
        if filter_by == 'name':
            query = Q(product__name__icontains=search_term)
        elif filter_by == 'type':
            query = Q(product__type__icontains=search_term)
        elif filter_by == 'price':
            query = Q(price__icontains=search_term)
        else:
            # Recherche globale sur nom ou type si pas de filtre
            query = Q(product__name__icontains=search_term) | Q(product__type__icontains=search_term)

    disponibilites = ProductAvailability.objects.select_related('product', 'form', 'pharmacy')\
                                                .filter(query)\
                                                .order_by('-id')

    # Pagination (5 éléments par page)
    paginator = Paginator(disponibilites, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "admin/produits.html", {
        "page_obj": page_obj,
        "search_term": search_term,
        "filter_by": filter_by,
    })
############ Vue pour supprimer un produit  ############
@login_required
@admin_required
def delete_product(request, id):
    produit = get_object_or_404(Product, id=id)
    pharmacy = produit.created_by

    subject = "Suppression d’un produit - PharmaConnect"
    html_message = f"""
    <div style="font-family: Arial; padding: 20px;">
        <h2 style="color: #dc3545;">Produit supprimé</h2>
        <p>Bonjour <strong>{pharmacy.pharmacy_name}</strong>,</p>
        <p>Le produit <strong style="color:#007bff;">{produit.name}</strong> a été <strong>supprimé</strong> par l'administration.</p>
        <p>Motifs possibles :</p>
        <ul>
            <li>Produit mal décrit ou erroné</li>
            <li>Non conforme à la réglementation</li>
            <li>Date de péremption proche ou dépassée</li>
        </ul>
        <p>Pour plus de détails, contactez : <a href="mailto:pharmaconnect200@gmail.com">pharmaconnect200@gmail.com</a>.</p>
        <p><strong>L'équipe PharmaConnect</strong></p>
    </div>
    """
    send_pharmacy_notification(pharmacy, subject, html_message)

    produit.delete()
    messages.success(request, "Produit supprimé avec succès.")
    return redirect('administration:manage_products')

############ Vue pour exporter le fichier excel  ############
@login_required
def export_products_excel(request):
    search = request.GET.get('search', '')
    filter_by = request.GET.get('filter_by', '')

    products = Product.objects.all()

    # Filtrage selon recherche et filtre choisi
    if search:
        if filter_by == 'name':
            products = products.filter(name__icontains=search)
        elif filter_by == 'type':
            products = products.filter(type__icontains=search)
        elif filter_by == 'price':
            # On filtre sur au moins une disponibilité avec ce prix
            products = products.filter(availabilities__price__icontains=search).distinct()
        else:
            # Recherche générale sur nom, type, description
            products = products.filter(
                Q(name__icontains=search) |
                Q(type__icontains=search) |
                Q(description__icontains=search)
            )

    products = products.order_by('-created_at')

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des produits"

    # En-tête colonnes
    ws.append([
        "Nom",
        "Type",
        "Description",
        "Mode d'utilisation",
        "Posologie",
        "Date d'expiration",
        "Date d'ajout",
        "Pharmacie",
        "Disponibilités (forme, dosage, prix)"
    ])

    # Remplissage des données
    for product in products:
        disponibilites = []
        for dispo in product.availabilities.all():
            dispo_str = f"{dispo.form.name} {dispo.dosage} - {dispo.price} FNG"
            disponibilites.append(dispo_str)
        disponibilites_text = "; ".join(disponibilites) if disponibilites else "Aucune"

        ws.append([
            product.name,
            product.type,
            product.description,
            product.usage_mode,
            product.posology,
            product.expiration_date.strftime("%d/%m/%Y") if product.expiration_date else '',
            product.created_at.strftime("%d/%m/%Y"),
            product.created_by.pharmacy_name if product.created_by else '',
            disponibilites_text,
        ])

    # Préparation de la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=produits.xlsx'

    wb.save(response)
    return response

############ Vue pour ajouter une nouvelle forme ############
@login_required
@admin_required
def manage_forms(request):
    forms_qs = Form.objects.all().order_by('name')  # Trie par nom
    paginator = Paginator(forms_qs, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    edit_id = request.GET.get('edit_id')
    form_instance = None
    if edit_id:
        form_instance = get_object_or_404(Form, id=edit_id)

    if request.method == 'POST':
        post_edit_id = request.POST.get('edit_id')
        if post_edit_id:
            form_instance = get_object_or_404(Form, id=post_edit_id)
            form = FormForm(request.POST, instance=form_instance)
        else:
            form = FormForm(request.POST)

        if form.is_valid():
            name = form.cleaned_data['name'].strip().lower()
            qs = Form.objects.filter(name__iexact=name)
            if form_instance:
                qs = qs.exclude(id=form_instance.id)
            if qs.exists():
                messages.error(request, "Cette forme pharmaceutique existe déjà.")
            else:
                form.save()
                messages.success(request, "Forme modifiée avec succès." if form_instance else "Forme ajoutée avec succès.")
                return redirect('administration:manage_forms')
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = FormForm(instance=form_instance)

    return render(request, "admin/formes.html", {
        "page_obj": page_obj,
        "form": form,
        "is_edit": bool(edit_id),
        "edit_id": edit_id,
    })

############ Vue pour supprimer une forme ############
@login_required
@admin_required
def delete_form(request, id):
    form = get_object_or_404(Form, id=id)
    form.delete()
    messages.success(request, "Forme pharmaceutique supprimée avec succès.")
    return redirect('administration:manage_forms')

############ Vue pour lister les pharmaciens abonnés ############
@login_required
@admin_required
def manage_subscriptions(request):
    abonnements_list = Abonnement.objects.select_related("user__pharmacy", "formule").order_by('-date_debut')
    paginator = Paginator(abonnements_list, 5)  # 5 abonnements par page
    page = request.GET.get("page")
    abonnements = paginator.get_page(page)
    return render(request, "admin/abonnements.html", {"abonnements": abonnements})

############ Vue pour activer ou desactiver un abonnement ############
@login_required
@admin_required
def toggle_abonnement(request, id):
    abonnement = get_object_or_404(Abonnement, id=id)
    abonnement.est_actif = not abonnement.est_actif
    abonnement.save()
    statut = "activé" if abonnement.est_actif else "désactivé"

    # Trouver la pharmacie associée à l'utilisateur de l'abonnement
    try:
        pharmacy = Pharmacy.objects.get(user=abonnement.user)
    except Pharmacy.DoesNotExist:
        pharmacy = None

    # Si abonnement désactivé, envoie un mail à la pharmacie
    if not abonnement.est_actif and pharmacy:
        subject = "Désactivation de votre abonnement - PharmaConnect"
        html_message = f"""
        <div style="font-family: Arial, sans-serif; padding: 20px; color: #333;">
            <h2 style="color: #dc3545;">Abonnement désactivé</h2>
            <p>Bonjour <strong>{pharmacy.pharmacy_name}</strong>,</p>
            <p>Votre abonnement a été <strong>désactivé</strong> sur <strong>PharmaConnect</strong>.</p>
            <p>Les raisons possibles :</p>
            <ul>
                <li>Expiration de la période d’abonnement</li>
                <li>Non-respect des conditions d'utilisation</li>
                <li>Demande de l'administration</li>
            </ul>
            <p>Pour toute question ou pour renouveler votre abonnement, merci de contacter notre support : 
               <a href="mailto:pharmaconnect200@gmail.com">pharmaconnect200@gmail.com</a>.</p>
            <p><strong>L’équipe PharmaConnect</strong></p>
        </div>
        """
        send_mail(
            subject,
            '',
            settings.DEFAULT_FROM_EMAIL,
            [pharmacy.email],
            html_message=html_message,
            fail_silently=True,
        )

    messages.success(request, f"L'abonnement a été {statut} avec succès.")
    return redirect('administration:manage_subscriptions')

############ Vue pour exporter la liste des abonnés en excel ############
@login_required
@admin_required
def export_abonnements_excel(request):
    abonnements = Abonnement.objects.select_related('user', 'formule')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Abonnements"
    ws.append([
        "Utilisateur", "Email", "Formule", "Prix", "Date Début", "Date Fin", "Statut"
    ])

    for ab in abonnements:
        ws.append([
            ab.user.get_full_name(),
            ab.user.email,
            ab.formule.nom,
            f"{ab.formule.prix} GNF",
            ab.date_debut.strftime("%d/%m/%Y %H:%M"),
            ab.date_expiration.strftime("%d/%m/%Y %H:%M") if ab.date_expiration else "",
            "Actif" if ab.est_valide else "Inactif"
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="abonnements.xlsx"'
    wb.save(response)
    return response

############ Vue pour afficher tous les produits expirés ############
@login_required
@admin_required
def expired_products(request):
    query = request.GET.get('q', '')

    produits = Product.objects.filter(expiration_date__lt=timezone.now())

    if query:
        produits = produits.filter(
            Q(name__icontains=query) | Q(created_by__pharmacy_name__icontains=query)
        )
    
    paginator = Paginator(produits.order_by('expiration_date'), 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "admin/produits_expirés.html", {
        "produits": page_obj,
        "query": query
    })

############ Vue pour exporter la liste des produits perimés en excel ############
@login_required
@admin_required
def export_expired_products_excel(request):
    produits = Product.objects.filter(expiration_date__lt=timezone.now()).order_by('expiration_date')

    # Création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits Expirés"

    # Entêtes
    headers = ["Nom du Produit", "Pharmacie", "Date d'Expiration"]
    ws.append(headers)

    # Remplir les données
    for produit in produits:
        ws.append([
            produit.name,
            produit.created_by.pharmacy_name,
            produit.expiration_date.strftime("%d/%m/%Y") if produit.expiration_date else ""
        ])

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # lettre colonne, ex 'A'
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=produits_expirés.xlsx'
    wb.save(response)
    return response
User = get_user_model()

####### Vue pour lister tous les utilisateurs ########
@login_required
@admin_required
def all_users(request):
    query = request.GET.get('q', '')
    users_qs = User.objects.filter(is_superuser=False)  # Exclure admin

    if query:
        users_qs = users_qs.filter(
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )

    paginator = Paginator(users_qs.order_by('-date_joined'), 5)
    page_number = request.GET.get('page')
    users_page = paginator.get_page(page_number)

    return render(request, 'admin/tous_les_utilisateurs.html', {
        'users': users_page,
        'query': query,
        'page_title': "Tous les utilisateurs",
    })

##### Vue pour bloquer un utilisateur et l'envoyer un email ####
@login_required
@admin_required
def block_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    # Basculer l'état actif/inactif
    user.is_active = not user.is_active
    user.save()

    # Déterminer si c'est un pharmacien
    try:
        pharmacy = user.pharmacy  # Si OneToOneField dans Pharmacy
    except Pharmacy.DoesNotExist:
        pharmacy = None

    if user.is_active:
        messages.success(request, f"L'utilisateur {user.get_full_name() or user.username} a été débloqué.")
    else:
        messages.warning(request, f"L'utilisateur {user.get_full_name() or user.username} a été bloqué.")

        # Si c'est un pharmacien, envoie d'email
        if pharmacy:
            subject = "Blocage de votre compte - PharmaConnect"
            html_message = f"""
                <div style="font-family: Arial; padding: 20px;">
                    <h2 style="color: #dc3545;">Compte bloqué</h2>
                    <p>Bonjour <strong>{pharmacy.pharmacy_name}</strong>,</p>
                    <p>Votre compte sur la plateforme <strong>PharmaConnect</strong> a été <strong style="color: red;">bloqué</strong> par l'administration.</p>
                    <p>Cette décision a été prise en raison d'une activité non conforme à nos règles, ou suite à des vérifications administratives.</p>
                    <p>Pour toute réclamation, veuillez nous contacter à <a href="pharmaconnect200@gmail.com">pharmaconnect200@gmail.com</a></p>
                    <p>Cordialement,<br><strong>L'équipe PharmaConnect</strong></p>
                </div>
            """
            print(f"ENVOI À : {pharmacy.email}")  # Debug
            send_pharmacy_notification(pharmacy, subject, html_message)

    # Rediriger selon l'origine
    if pharmacy:
        return redirect("administration:all_users")
    return redirect("administration:all_users")

##### Vue pour debloquer un utilisateur et l'envoyer un email ####
@login_required
@admin_required
def unblock_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if not user.is_active:
        user.is_active = True
        user.save()

        messages.success(request, f"L'utilisateur {user.get_full_name() or user.username} a été débloqué.")

        # Vérifier si utilisateur est pharmacien
        try:
            pharmacy = user.pharmacy
        except Pharmacy.DoesNotExist:
            pharmacy = None

        if pharmacy:
            subject = "Déblocage de votre compte - PharmaConnect"
            html_message = f"""
                <div style="font-family: Arial; padding: 20px;">
                    <h2 style="color: #198754;">Compte débloqué</h2>
                    <p>Bonjour <strong>{pharmacy.pharmacy_name}</strong>,</p>
                    <p>Bonne nouvelle ! Votre compte <strong>PharmaConnect</strong> a été <strong style="color: green;">débloqué</strong>.</p>
                    <p>Vous pouvez maintenant vous reconnecter à la plateforme et continuer vos activités.</p>
                    <p>Merci pour votre patience.</p>
                    <p>Cordialement,<br><strong>L'équipe PharmaConnect</strong></p>
                </div>
            """
            send_pharmacy_notification(pharmacy, subject, html_message)
    else:
        messages.info(request, f"L'utilisateur {user.get_full_name() or user.username} est déjà actif.")

    return redirect("administration:all_users")

############ Vue pour exporter la liste des utilisateurs bloqués en excel ############
@login_required
@admin_required
def export_blocked_users_excel(request):
    query = request.GET.get('q', '')
    users_qs = User.objects.filter(is_active=False).exclude(is_superuser=True)
    if query:
        users_qs = users_qs.filter(
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )
    users = users_qs.order_by('-date_joined')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs Bloqués"

    # Entêtes colonnes
    ws.append(["Nom", "Email", "Date d'inscription"])

    # Données
    for user in users:
        nom = user.get_full_name() or user.email
        ws.append([nom, user.email, user.date_joined.strftime("%d/%m/%Y %H:%M")])

    # Préparer la réponse HTTP avec fichier Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=utilisateurs_bloques.xlsx'
    wb.save(response)
    return response

##### Vue pour la statistique globale #######
@login_required
@admin_required
def statistics(request):
    # Générer les 6 derniers mois (par exemple)
    current_month = timezone.now().replace(day=1)
    months = [(current_month - timezone.timedelta(days=30 * i)).strftime("%B") for i in reversed(range(6))]

    # Abonnements par mois
    abonnements = (
        Abonnement.objects.annotate(month=TruncMonth("date_debut"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )

    # Pharmacies approuvées par mois
    pharmacies = (
        Pharmacy.objects.filter(is_approved=True)
        .annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )

    # Revenus mensuels : somme des prix des abonnements valides
    revenus = (
        Abonnement.objects.filter(est_actif=True, date_expiration__gte=now())
        .annotate(month=TruncMonth("date_debut"))
        .values("month")
        .annotate(total=Sum("formule__prix"))
        .order_by("month")
    )

    # Helper pour formatter les résultats
    def get_month_data(queryset):
        data_dict = {entry["month"].strftime("%B"): entry["total"] for entry in queryset}
        return [data_dict.get(month, 0) for month in months]

    def get_month_revenus(queryset):
        data_dict = {entry["month"].strftime("%B"): entry["total"] for entry in queryset}
        return [float(data_dict.get(month, 0) or 0) for month in months]

    context = {
        "months": json.dumps(months),
        "abonnements_per_month": json.dumps(get_month_data(abonnements)),
        "pharmacies_per_month": json.dumps(get_month_data(pharmacies)),
        "revenus_per_month": json.dumps(get_month_revenus(revenus)),
    }

    return render(request, "admin/statistiques.html", context)

##### Vue pour afficher la page profile #######
@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_profile(request):
    user = request.user
    return render(request, 'admin/admin_profile.html', {
        'user': user,
    })